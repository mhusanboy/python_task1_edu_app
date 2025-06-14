import csv
import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Helper for SQL escaping

def _sql_escape(value):
    if value is None:
        return "NULL"
    if isinstance(value, str):
        v = value.replace("'", "''")
        return f"'{v}'"
    return str(value)

def export_to_xlsx(platform_instance, filename="eduplatform_data.xlsx"):
    if not platform_instance.validate_data_for_export():
        print("Export cancelled due to data validation errors.")
        return

    wb = Workbook()

    ws_users = wb.active
    ws_users.title = "Users"
    user_headers = ["id", "full_name", "email", "role", "created_at", "phone", "address"]
    ws_users.append(user_headers)
    for user in platform_instance.users.values():
        ws_users.append([
            user._id, user._full_name, user._email, user.role.value, user._created_at,
            user.phone, user.address
        ])

    ws_students = wb.create_sheet("Students")
    student_headers = ["user_id", "full_name", "grade", "subjects", "assignments", "grades_data"]
    ws_students.append(student_headers)
    for student in platform_instance.students.values():
        ws_students.append([
            student._id, student._full_name, student.grade,
            str(student.subjects),
            str(student.assignments),
            str(student.grades)
        ])
    
    ws_teachers = wb.create_sheet("Teachers")
    teacher_headers = ["user_id", "full_name", "subjects", "classes", "workload"]
    ws_teachers.append(teacher_headers)
    for teacher in platform_instance.teachers.values():
        ws_teachers.append([
            teacher._id, teacher._full_name,
            ", ".join(teacher.subjects),
            ", ".join(teacher.classes),
            teacher.workload
        ])

    ws_parents = wb.create_sheet("Parents")
    parent_headers = ["user_id", "full_name", "children_ids", "notification_preferences"]
    ws_parents.append(parent_headers)
    for parent in platform_instance.parents.values():
        ws_parents.append([
            parent._id, parent._full_name,
            ", ".join(map(str, parent.children)),
            str(parent.notification_preferences)
        ])

    ws_assignments = wb.create_sheet("Assignments")
    assignment_headers = ["id", "title", "description", "deadline", "subject", "teacher_id", "class_id", "difficulty", "submissions_count", "grades_count"]
    ws_assignments.append(assignment_headers)
    for assignment in platform_instance.assignments.values():
        ws_assignments.append([
            assignment.id, assignment.title, assignment.description, assignment.deadline,
            assignment.subject, assignment.teacher_id, assignment.class_id,
            assignment.difficulty.value,
            len(assignment.submissions), len(assignment.grades)
        ])

    ws_grades = wb.create_sheet("Grades")
    grade_headers = ["id", "student_id", "subject", "value", "date", "teacher_id", "comment"]
    ws_grades.append(grade_headers)
    for grade in platform_instance.grades.values():
        ws_grades.append([
            grade.id, grade.student_id, grade.subject, grade.value,
            grade.date, grade.teacher_id, grade.comment
        ])

    ws_schedules = wb.create_sheet("Schedules")
    schedule_headers = ["id", "class_id", "day", "lessons_json"]
    ws_schedules.append(schedule_headers)
    for schedule in platform_instance.schedules.values():
        ws_schedules.append([
            schedule.id, schedule.class_id, schedule.day,
            str(schedule.lessons)
        ])
    
    ws_notifications = wb.create_sheet("Notifications")
    notification_headers = ["id", "message", "recipient_id", "created_at", "is_read", "priority"]
    ws_notifications.append(notification_headers)
    all_notifications = list(platform_instance.notifications.values())
    for user in platform_instance.users.values():
        for notif in user._notifications:
            if notif.id not in platform_instance.notifications:
                all_notifications.append(notif)
    
    for notif in all_notifications:
        ws_notifications.append([
            notif.id, notif.message, notif.recipient_id,
            notif.created_at, notif.is_read, notif.priority
        ])

    wb.save(filename)
    print(f"Data exported to {filename} successfully.")
    platform_instance.export_log.append({
        "timestamp": datetime.datetime.now().isoformat(),
        "action": "Manual XLSX Export",
        "filename": filename
    })

def export_to_csv(platform_instance, filename_prefix="eduplatform_data_"):
    if not platform_instance.validate_data_for_export():
        print("Export cancelled due to data validation errors.")
        return

    with open(f"{filename_prefix}users.csv", 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(["id", "full_name", "email", "role", "created_at", "phone", "address"])
        for user in platform_instance.users.values():
            writer.writerow([user._id, user._full_name, user._email, user.role.value, user._created_at, user.phone, user.address])
    print(f"Exported {filename_prefix}users.csv")

    with open(f"{filename_prefix}students.csv", 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(["user_id", "full_name", "grade", "subjects", "assignments", "grades_data"])
        for student in platform_instance.students.values():
            writer.writerow([student._id, student._full_name, student.grade, str(student.subjects), str(student.assignments), str(student.grades)])
    print(f"Exported {filename_prefix}students.csv")

    with open(f"{filename_prefix}teachers.csv", 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(["user_id", "full_name", "subjects", "classes", "workload"])
        for teacher in platform_instance.teachers.values():
            writer.writerow([teacher._id, teacher._full_name, ", ".join(teacher.subjects), ", ".join(teacher.classes), teacher.workload])
    print(f"Exported {filename_prefix}teachers.csv")

    with open(f"{filename_prefix}parents.csv", 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(["user_id", "full_name", "children_ids", "notification_preferences"])
        for parent in platform_instance.parents.values():
            writer.writerow([parent._id, parent._full_name, ", ".join(map(str, parent.children)), str(parent.notification_preferences)])
    print(f"Exported {filename_prefix}parents.csv")

    with open(f"{filename_prefix}assignments.csv", 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(["id", "title", "description", "deadline", "subject", "teacher_id", "class_id", "difficulty", "submissions_count", "grades_count"])
        for assignment in platform_instance.assignments.values():
            writer.writerow([assignment.id, assignment.title, assignment.description, assignment.deadline, assignment.subject, assignment.teacher_id, assignment.class_id, assignment.difficulty.value, len(assignment.submissions), len(assignment.grades)])
    print(f"Exported {filename_prefix}assignments.csv")

    with open(f"{filename_prefix}grades.csv", 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(["id", "student_id", "subject", "value", "date", "teacher_id", "comment"])
        for grade in platform_instance.grades.values():
            writer.writerow([grade.id, grade.student_id, grade.subject, grade.value, grade.date, grade.teacher_id, grade.comment])
    print(f"Exported {filename_prefix}grades.csv")

    with open(f"{filename_prefix}schedules.csv", 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(["id", "class_id", "day", "lessons_json"])
        for schedule in platform_instance.schedules.values():
            writer.writerow([schedule.id, schedule.class_id, schedule.day, str(schedule.lessons)])
    print(f"Exported {filename_prefix}schedules.csv")

    with open(f"{filename_prefix}notifications.csv", 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(["id", "message", "recipient_id", "created_at", "is_read", "priority"])
        all_notifications = list(platform_instance.notifications.values())
        for user in platform_instance.users.values():
            for notif in user._notifications:
                if notif.id not in platform_instance.notifications:
                    all_notifications.append(notif)
        for notif in all_notifications:
            writer.writerow([notif.id, notif.message, notif.recipient_id, notif.created_at, notif.is_read, notif.priority])
    print(f"Exported {filename_prefix}notifications.csv")

    print("All data exported to CSV successfully.")
    platform_instance.export_log.append({
        "timestamp": datetime.datetime.now().isoformat(),
        "action": "Manual CSV Export",
        "filename_prefix": filename_prefix
    })

def export_to_sql(platform_instance, filename="eduplatform_data.sql"):
    if not platform_instance.validate_data_for_export():
        print("Export cancelled due to data validation errors.")
        return

    sql_statements = []

    sql_statements.append("""
CREATE TABLE IF NOT EXISTS Users (
    id INT PRIMARY KEY,
    full_name VARCHAR(255) NOT NULL,
    email VARCHAR(255) UNIQUE NOT NULL,
    password_hash VARCHAR(255) NOT NULL,
    role VARCHAR(50) NOT NULL,
    created_at VARCHAR(255) NOT NULL,
    phone VARCHAR(20),
    address TEXT
);
""")

    sql_statements.append("""
CREATE TABLE IF NOT EXISTS Students (
    user_id INT PRIMARY KEY,
    grade VARCHAR(10) NOT NULL,
    subjects TEXT,
    assignments TEXT,
    grades_data TEXT,
    FOREIGN KEY (user_id) REFERENCES Users(id) ON DELETE CASCADE
);
""")

    sql_statements.append("""
CREATE TABLE IF NOT EXISTS Teachers (
    user_id INT PRIMARY KEY,
    subjects TEXT,
    classes TEXT,
    workload INT,
    FOREIGN KEY (user_id) REFERENCES Users(id) ON DELETE CASCADE
);
""")

    sql_statements.append("""
CREATE TABLE IF NOT EXISTS Parents (
    user_id INT PRIMARY KEY,
    children_ids TEXT,
    notification_preferences TEXT,
    FOREIGN KEY (user_id) REFERENCES Users(id) ON DELETE CASCADE
);
""")

    sql_statements.append("""
CREATE TABLE IF NOT EXISTS Assignments (
    id INT PRIMARY KEY,
    title VARCHAR(255) NOT NULL,
    description TEXT,
    deadline VARCHAR(255) NOT NULL,
    subject VARCHAR(100) NOT NULL,
    teacher_id INT NOT NULL,
    class_id VARCHAR(10) NOT NULL,
    difficulty VARCHAR(50) NOT NULL,
    submissions TEXT,
    grades TEXT,
    FOREIGN KEY (teacher_id) REFERENCES Users(id) ON DELETE CASCADE
);
""")

    sql_statements.append("""
CREATE TABLE IF NOT EXISTS Grades (
    id INT PRIMARY KEY,
    student_id INT NOT NULL,
    subject VARCHAR(100) NOT NULL,
    value INT NOT NULL CHECK (value >= 1 AND value <= 5),
    date VARCHAR(255) NOT NULL,
    teacher_id INT NOT NULL,
    comment TEXT,
    FOREIGN KEY (student_id) REFERENCES Users(id) ON DELETE CASCADE,
    FOREIGN KEY (teacher_id) REFERENCES Users(id) ON DELETE CASCADE
);
""")

    sql_statements.append("""
CREATE TABLE IF NOT EXISTS Schedules (
    id INT PRIMARY KEY,
    class_id VARCHAR(10) NOT NULL,
    day VARCHAR(20) NOT NULL,
    lessons TEXT
);
""")

    sql_statements.append("""
CREATE TABLE IF NOT EXISTS Notifications (
    id INT PRIMARY KEY,
    message TEXT NOT NULL,
    recipient_id INT NOT NULL,
    created_at VARCHAR(255) NOT NULL,
    is_read BOOLEAN NOT NULL,
    priority INT NOT NULL,
    FOREIGN KEY (recipient_id) REFERENCES Users(id) ON DELETE CASCADE
);
""")

    for user in platform_instance.users.values():
        sql_statements.append(f"""
INSERT INTO Users (id, full_name, email, password_hash, role, created_at, phone, address)
VALUES ({user._id}, {_sql_escape(user._full_name)}, {_sql_escape(user._email)}, {_sql_escape(user._password_hash)}, {_sql_escape(user.role.value)}, {_sql_escape(user._created_at)}, {_sql_escape(user.phone)}, {_sql_escape(user.address)});
""")

    for student in platform_instance.students.values():
        sql_statements.append(f"""
INSERT INTO Students (user_id, full_name, grade, subjects, assignments, grades_data)
VALUES ({student._id}, {_sql_escape(student._full_name)}, {_sql_escape(student.grade)}, {_sql_escape(str(student.subjects))}, {_sql_escape(str(student.assignments))}, {_sql_escape(str(student.grades))});
""")

    for teacher in platform_instance.teachers.values():
        sql_statements.append(f"""
INSERT INTO Teachers (user_id, full_name, subjects, classes, workload)
VALUES ({teacher._id}, {_sql_escape(teacher._full_name)}, {_sql_escape(", ".join(teacher.subjects))}, {_sql_escape(", ".join(teacher.classes))}, {teacher.workload});
""")
    
    for parent in platform_instance.parents.values():
        sql_statements.append(f"""
INSERT INTO Parents (user_id, full_name, children_ids, notification_preferences)
VALUES ({parent._id}, {_sql_escape(parent._full_name)}, {_sql_escape(", ".join(map(str, parent.children)))}, {_sql_escape(str(parent.notification_preferences))});
""")

    for assignment in platform_instance.assignments.values():
        sql_statements.append(f"""
INSERT INTO Assignments (id, title, description, deadline, subject, teacher_id, class_id, difficulty, submissions, grades)
VALUES ({assignment.id}, {_sql_escape(assignment.title)}, {_sql_escape(assignment.description)}, {_sql_escape(assignment.deadline)}, {_sql_escape(assignment.subject)}, {assignment.teacher_id}, {_sql_escape(assignment.class_id)}, {_sql_escape(assignment.difficulty.value)}, {_sql_escape(str(assignment.submissions))}, {_sql_escape(str(assignment.grades))});
""")

    for grade in platform_instance.grades.values():
        sql_statements.append(f"""
INSERT INTO Grades (id, student_id, subject, value, date, teacher_id, comment)
VALUES ({grade.id}, {grade.student_id}, {_sql_escape(grade.subject)}, {grade.value}, {_sql_escape(grade.date)}, {grade.teacher_id}, {_sql_escape(grade.comment)});
""")

    for schedule in platform_instance.schedules.values():
        sql_statements.append(f"""
INSERT INTO Schedules (id, class_id, day, lessons)
VALUES ({schedule.id}, {_sql_escape(schedule.class_id)}, {_sql_escape(schedule.day)}, {_sql_escape(str(schedule.lessons))});
""")

    all_notifications = list(platform_instance.notifications.values())
    for user in platform_instance.users.values():
        for notif in user._notifications:
            if notif.id not in platform_instance.notifications:
                all_notifications.append(notif)
    
    for notif in all_notifications:
        sql_statements.append(f"""
INSERT INTO Notifications (id, message, recipient_id, created_at, is_read, priority)
VALUES ({notif.id}, {_sql_escape(notif.message)}, {notif.recipient_id}, {_sql_escape(notif.created_at)}, {1 if notif.is_read else 0}, {notif.priority});
""")

    with open(filename, 'w', encoding='utf-8') as f:
        for statement in sql_statements:
            f.write(statement.strip() + "\n\n")

    print(f"SQL INSERT statements exported to {filename} successfully.")
    platform_instance.export_log.append({
        "timestamp": datetime.datetime.now().isoformat(),
        "action": "Manual SQL Export",
        "filename": filename
    })

def _scrape_data(url="https://www.olx.uz/"):
    print(f"\n--- Data Scraping Placeholder ---")
    print(f"Attempting to scrape data from {url}...")
    print("This is a placeholder. Actual web scraping is complex and requires:")
    print("1. Libraries like 'requests' for fetching content and 'BeautifulSoup' for parsing HTML.")
    print("2. Handling dynamic content (JavaScript rendering) often requires 'Selenium'.")
    print("3. Respecting website's robots.txt and terms of service.")
    print("4. Implementing robust error handling and retry mechanisms.")
    print("5. Data cleaning and structuring for scraped data.")
    print("For this in-memory application, live scraping is not implemented to keep the code simple.")
    print("--- End of Data Scraping Placeholder ---")
    return []